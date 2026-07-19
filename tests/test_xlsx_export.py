"""Tests for ``infrastructure.xlsx_export_service`` — safe local XLSX export.

Covers:
  - InventorySnapshot export
  - DailyAlertsSnapshot export
  - SupplierReorderResult export (assigned + unassigned sheets)
  - Greek worksheet names, headers, cell values, frozen rows
  - Deterministic row ordering
  - Invalid extension rejection
  - Existing-file non-overwrite protection
  - Failure cleanup (no partial output left behind)
  - Snapshot non-mutation guarantee
"""

from __future__ import annotations

import os
import sqlite3
import tempfile
from copy import deepcopy
from datetime import date, timedelta

import pytest
from openpyxl import load_workbook

from qt_app.data_source import (
    InventorySnapshot,
    InventoryProduct,
    DailyAlertsSnapshot,
    AlertItem,
    SupplierReorderResult,
    SupplierReorderGroup,
    ReorderCandidate,
    UnassignedReorderProduct,
)

from infrastructure.xlsx_export_service import (
    ExportResult,
    export_inventory_snapshot,
    export_daily_alerts,
    export_supplier_reorder,
)


# ── Snapshot factories ────────────────────────────────────────────────

def _inv_snapshot(*products: InventoryProduct) -> InventorySnapshot:
    return InventorySnapshot(
        total_matching=len(products),
        page=1, page_size=50,
        products=products,
    )


def _alerts_snapshot(*items: AlertItem) -> DailyAlertsSnapshot:
    return DailyAlertsSnapshot(
        low_stock_count=0, expiring_soon_count=0,
        expired_count=0, total_alerts=len(items),
        page=1, page_size=50, items=items,
    )


def _reorder_result(
    groups: tuple[SupplierReorderGroup, ...] = (),
    unassigned: tuple[UnassignedReorderProduct, ...] = (),
) -> SupplierReorderResult:
    return SupplierReorderResult.success(groups, unassigned)


# ── Helpers ───────────────────────────────────────────────────────────

def _rows(ws) -> list[list]:
    """Return all data rows (excluding header) as list of cell values."""
    return [
        [cell.value for cell in row]
        for row in ws.iter_rows(min_row=2, values_only=False)
    ]


def _temp_xlsx() -> str:
    """Return a path that does NOT exist yet."""
    fd, path = tempfile.mkstemp(suffix=".xlsx")
    os.close(fd)
    os.remove(path)  # must not exist for export
    return path


# ═══════════════════════════════════════════════════════════════════════
# Inventory export
# ═══════════════════════════════════════════════════════════════════════

class TestInventoryExport:

    def test_headers(self, tmp_path):
        snap = _inv_snapshot()
        path = str(tmp_path / "inv.xlsx")
        r = export_inventory_snapshot(snap, path)
        assert r.ok
        wb = load_workbook(path)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == [
            "Barcode", "Προϊόν", "Απόθεμα", "Ημ/νία Λήξης",
            "Τιμή", "Προμηθευτής", "Κατάσταση",
        ]

    def test_worksheet_name(self, tmp_path):
        snap = _inv_snapshot()
        path = str(tmp_path / "inv.xlsx")
        export_inventory_snapshot(snap, path)
        wb = load_workbook(path)
        assert wb.active.title == "Αποθήκη"

    def test_frozen_header(self, tmp_path):
        snap = _inv_snapshot()
        path = str(tmp_path / "inv.xlsx")
        export_inventory_snapshot(snap, path)
        wb = load_workbook(path)
        assert wb.active.freeze_panes == "A2"

    def test_empty_snapshot_yields_only_header(self, tmp_path):
        snap = _inv_snapshot()
        path = str(tmp_path / "inv.xlsx")
        r = export_inventory_snapshot(snap, path)
        assert r.ok
        wb = load_workbook(path)
        ws = wb.active
        assert ws.max_row == 1  # header only
        assert ws.max_column == 7

    def test_single_product_values(self, tmp_path):
        prod = InventoryProduct(
            barcode="5200000000017", name="Παρακεταμόλη",
            stock=42, expiry_date="2026-12-31", price=3.50,
            supplier_id=1, supplier_name="Demo Pharma",
            status_labels=("Ληγμένο", "Χαμηλό απόθεμα"),
        )
        snap = _inv_snapshot(prod)
        path = str(tmp_path / "inv.xlsx")
        r = export_inventory_snapshot(snap, path)
        assert r.ok
        wb = load_workbook(path)
        data = _rows(wb.active)
        assert len(data) == 1
        row = data[0]
        assert row[0] == "5200000000017"
        assert row[1] == "Παρακεταμόλη"
        assert row[2] == 42
        assert row[3] == "2026-12-31"
        assert row[4] == 3.50
        assert row[5] == "Demo Pharma"
        assert row[6] == "Ληγμένο, Χαμηλό απόθεμα"

    def test_missing_expiry_preserved(self, tmp_path):
        prod = InventoryProduct(
            barcode="X", name="Test", stock=1,
            expiry_date="—", price=0.0,
            supplier_id=None, supplier_name="—",
            status_labels=("—",),
        )
        snap = _inv_snapshot(prod)
        path = str(tmp_path / "inv.xlsx")
        export_inventory_snapshot(snap, path)
        wb = load_workbook(path)
        row = _rows(wb.active)[0]
        assert row[3] == "—"

    def test_deterministic_ordering(self, tmp_path):
        a = InventoryProduct("A", "Άλφα", 1, "—", 0, None, "—", ("—",))
        b = InventoryProduct("B", "Βήτα", 1, "—", 0, None, "—", ("—",))
        c = InventoryProduct("C", "Γάμμα", 1, "—", 0, None, "—", ("—",))
        snap = _inv_snapshot(a, b, c)
        path = str(tmp_path / "inv.xlsx")
        export_inventory_snapshot(snap, path)
        wb = load_workbook(path)
        data = _rows(wb.active)
        assert [r[1] for r in data] == ["Άλφα", "Βήτα", "Γάμμα"]

    def test_does_not_mutate_snapshot(self, tmp_path):
        prod = InventoryProduct("X", "Test", 10, "—", 5.0, 1, "S", ("—",))
        snap = _inv_snapshot(prod)
        snap_copy = deepcopy(snap)
        path = str(tmp_path / "inv.xlsx")
        export_inventory_snapshot(snap, path)
        assert snap == snap_copy


# ═══════════════════════════════════════════════════════════════════════
# Daily Alerts export
# ═══════════════════════════════════════════════════════════════════════

class TestDailyAlertsExport:

    def test_headers(self, tmp_path):
        snap = _alerts_snapshot()
        path = str(tmp_path / "alerts.xlsx")
        r = export_daily_alerts(snap, path)
        assert r.ok
        wb = load_workbook(path)
        ws = wb.active
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        assert headers == [
            "Barcode", "Προϊόν", "Απόθεμα", "Ημ/νία Λήξης",
            "Τιμή", "Λόγοι Ειδοποίησης",
        ]

    def test_worksheet_name(self, tmp_path):
        snap = _alerts_snapshot()
        path = str(tmp_path / "alerts.xlsx")
        export_daily_alerts(snap, path)
        wb = load_workbook(path)
        assert wb.active.title == "Ειδοποιήσεις"

    def test_frozen_header(self, tmp_path):
        snap = _alerts_snapshot()
        path = str(tmp_path / "alerts.xlsx")
        export_daily_alerts(snap, path)
        wb = load_workbook(path)
        assert wb.active.freeze_panes == "A2"

    def test_alert_with_greek_reasons(self, tmp_path):
        item = AlertItem(
            barcode="5200000000017", name="Παρακεταμόλη",
            stock=5, expiry_date="2026-07-20", price=3.50,
            reasons=("Ληγμένο", "Χαμηλό απόθεμα"),
        )
        snap = _alerts_snapshot(item)
        path = str(tmp_path / "alerts.xlsx")
        export_daily_alerts(snap, path)
        wb = load_workbook(path)
        data = _rows(wb.active)
        assert len(data) == 1
        row = data[0]
        assert row[0] == "5200000000017"
        assert row[1] == "Παρακεταμόλη"
        assert row[2] == 5
        assert row[3] == "2026-07-20"
        assert row[4] == 3.50
        assert row[5] == "Ληγμένο, Χαμηλό απόθεμα"

    def test_expiring_soon_reason(self, tmp_path):
        item = AlertItem(
            barcode="Z", name="Zeta", stock=20,
            expiry_date="2026-08-01", price=2.00,
            reasons=("Λήγει σύντομα (2026-08-01)",),
        )
        snap = _alerts_snapshot(item)
        path = str(tmp_path / "alerts.xlsx")
        export_daily_alerts(snap, path)
        wb = load_workbook(path)
        row = _rows(wb.active)[0]
        assert row[5] == "Λήγει σύντομα (2026-08-01)"

    def test_deterministic_ordering(self, tmp_path):
        a = AlertItem("A", "Άλφα", 1, "—", 0, ("—",))
        b = AlertItem("B", "Βήτα", 1, "—", 0, ("—",))
        snap = _alerts_snapshot(a, b)
        path = str(tmp_path / "alerts.xlsx")
        export_daily_alerts(snap, path)
        wb = load_workbook(path)
        data = _rows(wb.active)
        assert [r[1] for r in data] == ["Άλφα", "Βήτα"]

    def test_does_not_mutate_snapshot(self, tmp_path):
        item = AlertItem("X", "Test", 10, "—", 5.0, ("—",))
        snap = _alerts_snapshot(item)
        snap_copy = deepcopy(snap)
        path = str(tmp_path / "alerts.xlsx")
        export_daily_alerts(snap, path)
        assert snap == snap_copy


# ═══════════════════════════════════════════════════════════════════════
# Supplier Reorder export
# ═══════════════════════════════════════════════════════════════════════

class TestSupplierReorderExport:

    def test_sheet_names(self, tmp_path):
        r = _reorder_result()
        path = str(tmp_path / "reorder.xlsx")
        export_supplier_reorder(r, path)
        wb = load_workbook(path)
        assert wb.sheetnames == [
            "Υποψήφιοι Αναπαραγγελίας",
            "Προϊόντα Χωρίς Προμηθευτή",
        ]

    def test_candidates_sheet_has_supplier_sections(self, tmp_path):
        prod = ReorderCandidate(
            barcode="5200000000017", name="Παρακεταμόλη",
            stock=3, threshold=10, expiry_date="2026-12-31", price=3.50,
        )
        group = SupplierReorderGroup(
            supplier_id=1, supplier_name="Demo Pharma",
            products=(prod,),
        )
        r = _reorder_result(groups=(group,))
        path = str(tmp_path / "reorder.xlsx")
        export_supplier_reorder(r, path)
        wb = load_workbook(path)
        ws = wb["Υποψήφιοι Αναπαραγγελίας"]

        # Row 1: supplier header
        assert ws.cell(row=1, column=1).value == "Προμηθευτής: Demo Pharma"
        # Row 2: column headers
        prod_headers = [
            ws.cell(row=2, column=c).value for c in range(1, 8)
        ]
        assert prod_headers == [
            "Barcode", "Προϊόν", "Απόθεμα", "Όριο",
            "Ημ/νία Λήξης", "Τιμή", "Προμηθευτής",
        ]
        # Row 3: product data
        assert ws.cell(row=3, column=1).value == "5200000000017"
        assert ws.cell(row=3, column=2).value == "Παρακεταμόλη"
        assert ws.cell(row=3, column=3).value == 3
        assert ws.cell(row=3, column=4).value == 10
        assert ws.cell(row=3, column=5).value == "2026-12-31"
        assert ws.cell(row=3, column=6).value == 3.50
        assert ws.cell(row=3, column=7).value == "Demo Pharma"

    def test_unassigned_sheet_headers_and_values(self, tmp_path):
        unassigned = UnassignedReorderProduct(
            barcode="NO_SUP", name="Orphan", stock=2,
            threshold=10, expiry_date="—", price=1.0,
            reason="Χωρίς προμηθευτή",
        )
        r = _reorder_result(unassigned=(unassigned,))
        path = str(tmp_path / "reorder.xlsx")
        export_supplier_reorder(r, path)
        wb = load_workbook(path)
        ws = wb["Προϊόντα Χωρίς Προμηθευτή"]

        headers = [ws.cell(row=1, column=c).value for c in range(1, 8)]
        assert headers == [
            "Barcode", "Προϊόν", "Απόθεμα", "Όριο",
            "Ημ/νία Λήξης", "Τιμή", "Λόγος",
        ]

        row = [ws.cell(row=2, column=c).value for c in range(1, 8)]
        assert row[0] == "NO_SUP"
        assert row[1] == "Orphan"
        assert row[2] == 2
        assert row[3] == 10
        assert row[4] == "—"
        assert row[5] == 1.0
        assert row[6] == "Χωρίς προμηθευτή"

    def test_missing_supplier_reason(self, tmp_path):
        unassigned = UnassignedReorderProduct(
            barcode="DELETED", name="Ghost", stock=1,
            threshold=5, expiry_date="—", price=0.5,
            reason="Ο προμηθευτής δεν υπάρχει",
        )
        r = _reorder_result(unassigned=(unassigned,))
        path = str(tmp_path / "reorder.xlsx")
        export_supplier_reorder(r, path)
        wb = load_workbook(path)
        ws = wb["Προϊόντα Χωρίς Προμηθευτή"]
        assert ws.cell(row=2, column=7).value == "Ο προμηθευτής δεν υπάρχει"

    def test_supplier_name_unambiguous_per_row(self, tmp_path):
        """Each product row must carry its owning supplier name."""
        a1 = ReorderCandidate("A", "Alpha", 1, 10, "—", 1.0)
        b1 = ReorderCandidate("B", "Beta", 1, 10, "—", 1.0)
        g1 = SupplierReorderGroup(1, "Supplier One", (a1,))
        g2 = SupplierReorderGroup(2, "Supplier Two", (b1,))
        r = _reorder_result(groups=(g1, g2))
        path = str(tmp_path / "reorder.xlsx")
        export_supplier_reorder(r, path)
        wb = load_workbook(path)
        ws = wb["Υποψήφιοι Αναπαραγγελίας"]

        # Find data rows (skip header/label rows that aren't product data)
        all_values = []
        for row in ws.iter_rows(min_row=1, values_only=True):
            all_values.append(row)

        # Row 0: Supplier One header
        # Row 1: product headers
        # Row 2: A/Alpha product row → supplier col should be "Supplier One"
        # Row 3: blank separator
        # Row 4: Supplier Two header
        # Row 5: product headers
        # Row 6: B/Beta product row → supplier col should be "Supplier Two"

        assert all_values[0][0] == "Προμηθευτής: Supplier One"
        assert all_values[2][0] == "A"  # barcode
        assert all_values[2][6] == "Supplier One"
        assert all_values[4][0] == "Προμηθευτής: Supplier Two"
        assert all_values[6][0] == "B"
        assert all_values[6][6] == "Supplier Two"

    def test_no_quantities_or_formulas(self, tmp_path):
        prod = ReorderCandidate("X", "Test", 3, 10, "—", 5.0)
        group = SupplierReorderGroup(1, "S", (prod,))
        r = _reorder_result(groups=(group,))
        path = str(tmp_path / "reorder.xlsx")
        export_supplier_reorder(r, path)
        wb = load_workbook(path)
        ws = wb["Υποψήφιοι Αναπαραγγελίας"]
        # Check no cell contains a formula
        for row in ws.iter_rows():
            for cell in row:
                assert not str(cell.value).startswith("="), \
                    f"Formula found at {cell.coordinate}: {cell.value}"

    def test_deterministic_ordering(self, tmp_path):
        a = ReorderCandidate("A", "Άλφα", 1, 10, "—", 1.0)
        b = ReorderCandidate("B", "Βήτα", 1, 10, "—", 2.0)
        c = ReorderCandidate("C", "Γάμμα", 1, 10, "—", 3.0)
        group = SupplierReorderGroup(1, "Supplier", (a, b, c))
        r = _reorder_result(groups=(group,))
        path = str(tmp_path / "reorder.xlsx")
        export_supplier_reorder(r, path)
        wb = load_workbook(path)
        ws = wb["Υποψήφιοι Αναπαραγγελίας"]
        # Product rows start at row 3 (after supplier header + column headers)
        names = [ws.cell(row=r, column=2).value for r in range(3, 6)]
        assert names == ["Άλφα", "Βήτα", "Γάμμα"]

    def test_does_not_mutate_result(self, tmp_path):
        prod = ReorderCandidate("X", "Test", 3, 10, "—", 5.0)
        group = SupplierReorderGroup(1, "S", (prod,))
        r = _reorder_result(groups=(group,))
        r_copy = deepcopy(r)
        path = str(tmp_path / "reorder.xlsx")
        export_supplier_reorder(r, path)
        assert r == r_copy


# ═══════════════════════════════════════════════════════════════════════
# Validation / safety
# ═══════════════════════════════════════════════════════════════════════

class TestValidation:

    def test_rejects_non_xlsx_extension(self, tmp_path):
        snap = _inv_snapshot()
        path = str(tmp_path / "inv.xls")
        r = export_inventory_snapshot(snap, path)
        assert not r.ok
        assert "Μη έγκυρη διαδρομή" in r.error_message
        assert ".xlsx" in r.error_message
        assert not os.path.exists(path)

    def test_rejects_no_extension(self, tmp_path):
        snap = _inv_snapshot()
        path = str(tmp_path / "justafile")
        r = export_inventory_snapshot(snap, path)
        assert not r.ok
        assert not os.path.exists(path)

    def test_rejects_csv_extension(self, tmp_path):
        snap = _inv_snapshot()
        path = str(tmp_path / "out.csv")
        r = export_daily_alerts(snap, path)
        assert not r.ok
        assert not os.path.exists(path)

    def test_rejects_existing_file(self, tmp_path):
        """Export must fail when target already exists — no silent overwrite."""
        snap = _inv_snapshot()
        path = str(tmp_path / "existing.xlsx")
        # Create a placeholder file
        path.encode()  # ensure str
        with open(path, "wb") as f:
            f.write(b"not a real xlsx")
        r = export_inventory_snapshot(snap, path)
        assert not r.ok
        assert "υπάρχει ήδη" in r.error_message
        # Original content must be untouched
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0

    def test_rejects_existing_file_alerts(self, tmp_path):
        snap = _alerts_snapshot()
        path = str(tmp_path / "existing.xlsx")
        with open(path, "wb") as f:
            f.write(b"placeholder")
        r = export_daily_alerts(snap, path)
        assert not r.ok
        assert "υπάρχει ήδη" in r.error_message

    def test_rejects_existing_file_reorder(self, tmp_path):
        reorder_data = _reorder_result()
        path = str(tmp_path / "existing.xlsx")
        with open(path, "wb") as f:
            f.write(b"placeholder")
        r = export_supplier_reorder(reorder_data, path)
        assert not r.ok
        assert "υπάρχει ήδη" in r.error_message

    def test_error_path_is_none(self, tmp_path):
        snap = _inv_snapshot()
        path = str(tmp_path / "inv.csv")
        r = export_inventory_snapshot(snap, path)
        assert not r.ok
        assert r.path is None

    def test_success_path_is_set(self, tmp_path):
        snap = _inv_snapshot()
        path = str(tmp_path / "ok.xlsx")
        r = export_inventory_snapshot(snap, path)
        assert r.ok
        assert r.path == path

    def test_no_partial_output_on_write_failure(self, tmp_path):
        """If the target directory is read-only, no partial file remains."""
        snap = _inv_snapshot()
        # Path on a non-existent directory — write will fail
        bad_path = str(tmp_path / "nonexistent" / "out.xlsx")
        r = export_inventory_snapshot(snap, bad_path)
        assert not r.ok
        # The target file must not exist
        assert not os.path.exists(bad_path)


# ═══════════════════════════════════════════════════════════════════════
# No side effects beyond file creation
# ═══════════════════════════════════════════════════════════════════════

class TestNoSideEffects:

    def test_export_does_not_touch_database(self, tmp_path, db):
        """Export from a snapshot — db fixture is never accessed."""
        prod = InventoryProduct("X", "Test", 1, "—", 0, None, "—", ("—",))
        snap = _inv_snapshot(prod)
        path = str(tmp_path / "out.xlsx")
        # Just verify no crash — export only reads from the snapshot
        r = export_inventory_snapshot(snap, path)
        assert r.ok
        # db must still be usable
        assert db.get_all_config() is not None

    def test_import_safe_without_pyside6(self):
        """The export module must not require PySide6 at import time.
        (TYPE_CHECKING guards ensure this.)"""
        import infrastructure.xlsx_export_service as svc
        assert hasattr(svc, "export_inventory_snapshot")
        assert hasattr(svc, "export_daily_alerts")
        assert hasattr(svc, "export_supplier_reorder")
