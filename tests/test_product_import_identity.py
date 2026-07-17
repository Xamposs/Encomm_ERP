"""Tests for import source identity (fingerprinting + verification)."""

import hashlib, json, pytest, os, stat
import openpyxl
from infrastructure.product_import_identity import (
    fingerprint_import_source, verify_import_source,
    ImportSourceSignature, _sha256_mapping, CHUNK_SIZE,
)
from infrastructure.product_import_preview import ImportColumnMapping


M = ImportColumnMapping("Barcode", "Name", "Stock", "Price", "Expiry")
M2 = ImportColumnMapping("Code", "Product", "Qty", "Cost", "Exp")


def _xlsx(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(path)


class TestFingerprint:

    def test_deterministic(self, tmp_path):
        p = str(tmp_path / "a.xlsx")
        _xlsx(p, ["Barcode", "Name", "Stock", "Price", "Expiry"], [])
        s1 = fingerprint_import_source(p, M)
        s2 = fingerprint_import_source(p, M)
        assert s1 == s2

    def test_file_change_changes_hash(self, tmp_path):
        p = str(tmp_path / "a.xlsx")
        _xlsx(p, ["Barcode", "Name", "Stock", "Price", "Expiry"],
              [["A", "N", 1, 1.0, ""]])
        s1 = fingerprint_import_source(p, M)
        wb = openpyxl.load_workbook(p)
        ws = wb.active
        ws.cell(row=2, column=1, value="B")
        wb.save(p)
        s2 = fingerprint_import_source(p, M)
        assert s1.file_sha256 != s2.file_sha256

    def test_mapping_change_changes_hash(self, tmp_path):
        p = str(tmp_path / "a.xlsx")
        _xlsx(p, ["Barcode", "Name", "Stock", "Price", "Expiry"], [])
        s1 = fingerprint_import_source(p, M)
        s2 = fingerprint_import_source(p, M2)
        assert s1.mapping_sha256 != s2.mapping_sha256

    def test_json_mapping_no_collision(self):
        """Delimiter-join would collide: A|B + C collides with A + B|C."""
        m_a = ImportColumnMapping("A|B", "C", "D", "E", "F")
        m_b = ImportColumnMapping("A", "B|C", "D", "E", "F")
        assert _sha256_mapping(m_a) != _sha256_mapping(m_b)

    def test_directory_rejected(self, tmp_path):
        with pytest.raises(ValueError, match="κανονικό"):
            fingerprint_import_source(str(tmp_path), M)

    def test_format_version_mismatch_fails(self, tmp_path):
        p = str(tmp_path / "a.xlsx")
        _xlsx(p, ["Barcode", "Name", "Stock", "Price", "Expiry"], [])
        s = fingerprint_import_source(p, M)
        # Mutate format_version
        altered = ImportSourceSignature(
            format_version=999,
            file_size_bytes=s.file_size_bytes,
            file_sha256=s.file_sha256,
            mapping_sha256=s.mapping_sha256,
        )
        assert not verify_import_source(altered, p, M)

    def test_genuine_chunked_reads(self, tmp_path):
        """Monkeypatch to prove that file reads use chunks ≤ CHUNK_SIZE."""
        p = str(tmp_path / "big.bin")
        size = CHUNK_SIZE * 2 + 500
        with open(p, "wb") as f:
            f.write(os.urandom(size))
        reads = []

        orig_open = open
        def tracking_open(*a, **kw):
            f = orig_open(*a, **kw)
            orig_read = f.read
            def tracker(n=None):
                data = orig_read(n)
                reads.append(len(data))
                return data
            f.read = tracker
            return f

        import builtins
        builtins.open = tracking_open
        try:
            from infrastructure.product_import_identity import _sha256_file
            _sha256_file(p)
        finally:
            builtins.open = orig_open

        assert len(reads) >= 3, f"expected 3+ reads, got {reads}"
        for r in reads[:-1]:  # skip EOF read (0 bytes)
            assert 0 < r <= CHUNK_SIZE, f"chunk {r} > {CHUNK_SIZE}"

    def test_verify_match(self, tmp_path):
        p = str(tmp_path / "a.xlsx")
        _xlsx(p, ["Barcode", "Name", "Stock", "Price", "Expiry"], [])
        s = fingerprint_import_source(p, M)
        assert verify_import_source(s, p, M)

    def test_verify_mismatch(self, tmp_path):
        p = str(tmp_path / "a.xlsx")
        _xlsx(p, ["Barcode", "Name", "Stock", "Price", "Expiry"], [])
        s = fingerprint_import_source(p, M)
        assert not verify_import_source(s, p, M2)

    def test_no_write_sql(self):
        import inspect
        from infrastructure import product_import_identity as pid
        src = inspect.getsource(pid.fingerprint_import_source)
        for pat in ["INSERT", "UPDATE", "DELETE", "DROP",
                     "ALTER", "CREATE", "REPLACE", "sqlite3"]:
            assert pat not in src, f"Forbidden '{pat}' in identity"
