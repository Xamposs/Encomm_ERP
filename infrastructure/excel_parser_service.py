"""
Excel / CSV Parser Service – Supplier Price-List Ingestion

Reads supplier invoice files (.xlsx, .xls, .csv) and returns a clean
list of (barcode, name, stock, expiry_date, price) tuples ready for
the bulk_upsert_products database method.
"""

import csv
import logging
import os
from datetime import datetime

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


def _normalize_expiry(raw: str) -> str:
    """Normalize a raw expiry value to ``YYYY-MM-DD`` or ``""``.

    Accepts ``YYYY-MM-DD`` (optionally with a trailing time component) and
    the common European ``DD/MM/YYYY`` form. Empty/whitespace → ``""``.
    Raises ``ValueError`` for anything that cannot be normalized, so the
    caller (``_cast_row``) skips the row with a clear warning instead of
    letting a garbage date silently hide expired stock in the DB.
    """
    raw = (raw or "").strip()
    if not raw:
        return ""
    # Strip a trailing timestamp if present.
    candidate = raw.split(" ")[0]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(candidate, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    raise ValueError(f"Unparseable expiry date: {raw!r}")


class ExcelParserService:
    """Parses supplier files into normalised product tuples."""

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------
    def parse_supplier_file(self, file_path: str) -> list:
        """
        Parse a supplier invoice file and return product tuples.

        Parameters
        ----------
        file_path : str
            Absolute or relative path to a .csv / .xlsx / .xls file.

        Returns
        -------
        list[tuple]
            Each tuple: ``(barcode: str, name: str, stock: int,
                           expiry_date: str, price: float[, supplier_id: int])``
        """
        if not os.path.isfile(file_path):
            logger.error(f"File not found: {file_path}")
            return []

        ext = os.path.splitext(file_path)[1].lower()
        logger.info(f"Parsing supplier file: {file_path} ({ext})")

        if ext == ".csv":
            rows = self._read_csv(file_path)
        elif ext in (".xlsx", ".xls"):
            rows = self._read_excel(file_path)
        else:
            logger.error(f"Unsupported file extension: {ext}")
            return []

        products = []
        for idx, raw in enumerate(rows, start=1):
            parsed = self._cast_row(raw, idx)
            if parsed is not None:
                products.append(parsed)

        logger.info(
            f"Parsing complete. {len(products)} valid products extracted "
            f"from {file_path}"
        )
        return products

    # ------------------------------------------------------------------
    # Readers
    # ------------------------------------------------------------------
    def _read_csv(self, file_path: str):
        """Yield rows from a CSV file (UTF-8 with Greek fallback)."""
        for encoding in ("utf-8", "iso-8859-7"):
            try:
                with open(file_path, newline="", encoding=encoding) as fh:
                    reader = csv.reader(fh)
                    header = next(reader, None)  # skip header row
                    if header is not None:
                        logger.debug(f"CSV header ({encoding}): {header}")
                    return list(reader)
            except UnicodeDecodeError:
                logger.debug(f"CSV decode failed with {encoding}, trying next…")
                continue

        logger.error(f"Could not decode CSV file: {file_path}")
        return []

    def _read_excel(self, file_path: str):
        """Yield rows from an Excel workbook (.xlsx / .xls)."""
        wb = None
        try:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            header = next(rows_iter, None)  # skip header row
            if header is not None:
                logger.debug(f"Excel header: {header}")
            rows = list(rows_iter)
            return rows
        except Exception:
            logger.exception(f"Failed to read Excel file: {file_path}")
            return []
        finally:
            if wb is not None:
                wb.close()

    # ------------------------------------------------------------------
    # Row casting & validation
    # ------------------------------------------------------------------
    def _cast_row(self, raw: tuple, row_idx: int):
        """
        Convert a raw row to ``(barcode, name, stock, expiry_date, price)``.

        Returns *None* and logs a warning for any row that cannot be
        safely cast, so the rest of the file continues uninterrupted.
        """
        try:
            # Flatten any None padding and ensure at least 5 columns.
            # Note: supplier_id (6th column) is optional — the DB layer
            # handles it via COALESCE in bulk_upsert.
            values = list(raw) + [None] * 5
            values = values[:5]

            barcode = str(values[0]).strip()
            name = str(values[1]).strip()
            stock = int(float(values[2])) if values[2] is not None else 0
            raw_expiry = str(values[3]).strip() if values[3] is not None else ""
            price = round(float(values[4]), 2) if values[4] is not None else 0.0

            # Validate expiry so malformed dates never reach the DB (where they
            # would defeat SQLite date() comparisons and hide expired stock).
            expiry_date = _normalize_expiry(raw_expiry)

            if not barcode:
                raise ValueError("Empty barcode")

            return (barcode, name, stock, expiry_date, price)

        except (ValueError, TypeError) as exc:
            logger.warning(
                f"Skipping malformed row {row_idx}: {exc} | raw={raw}"
            )
            return None
