"""Safe local XLSX export service for ERP read models.

Receives already-loaded snapshot data from callers — never queries SQLite
directly, never writes to the database, never integrates with Qt/PySide6.

Export scopes:
  1. InventorySnapshot     → worksheet "Αποθήκη"
  2. DailyAlertsSnapshot   → worksheet "Ειδοποιήσεις"
  3. SupplierReorderResult → worksheets "Υποψήφιοι Αναπαραγγελίας"
                              + "Προϊόντα Χωρίς Προμηθευτή"
"""

from __future__ import annotations

import errno
import os
import tempfile
from dataclasses import dataclass
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from qt_app.data_source import (
        InventorySnapshot,
        DailyAlertsSnapshot,
        SupplierReorderResult,
    )


# ── Typed result contract ─────────────────────────────────────────────

@dataclass(frozen=True)
class ExportResult:
    """Carries either the written file path or a Greek error message."""
    ok: bool
    path: str | None = None
    error_message: str = ""

    @classmethod
    def success(cls, path: str) -> "ExportResult":
        return cls(ok=True, path=path)

    @classmethod
    def failure(cls, message: str) -> "ExportResult":
        return cls(ok=False, error_message=message)


# ── Internal helpers ──────────────────────────────────────────────────

def _validate_xlsx_path(path: str) -> ExportResult | None:
    """Return ExportResult.failure if *path* is not a valid .xlsx target,
    or None when validation passes."""
    if not path.lower().endswith(".xlsx"):
        return ExportResult.failure(
            f"Μη έγκυρη διαδρομή αρχείου: '{path}'. "
            "Η εξαγωγή υποστηρίζει μόνο αρχεία .xlsx."
        )
    if os.path.exists(path):
        return ExportResult.failure(
            f"Το αρχείο '{os.path.basename(path)}' υπάρχει ήδη. "
            "Η αντικατάσταση υπαρχόντων αρχείων δεν επιτρέπεται."
        )
    return None


def _freeze_header(ws) -> None:
    """Freeze the first row so headers remain visible on scroll."""
    ws.freeze_panes = "A2"


def _auto_width(ws, min_width: int = 10, max_width: int = 40) -> None:
    """Set column widths proportional to content length."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        width = max(min_width, min(max_len + 3, max_width))
        ws.column_dimensions[col_letter].width = width


def _write_header(ws, headers: list[str]) -> None:
    """Write bold header row."""
    bold = Font(bold=True)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold


def _safe_write(path: str, wb: Workbook) -> ExportResult:
    """Write workbook to a temp file, then create a hard link into
    final position.

    ``os.link(tmp_path, path)`` creates a new directory entry for the
    temp inode under *path* — it succeeds *only* when *path* does not
    yet exist.  Both POSIX and Windows raise ``FileExistsError`` when
    *dst* already exists, making this a genuinely non-overwriting
    finalization that works identically on every supported OS.

    After a successful link the temp directory entry is removed with
    ``os.unlink``; the data survives through *path*'s directory entry.

    On any failure the temp file is cleaned up.  The target path is
    **never** removed — if a concurrent process created it in the
    window between validation and finalization, ``os.link`` raises
    ``FileExistsError`` and the concurrent file survives untouched.

    If ``os.link`` is unavailable on the filesystem (cross-device
    link, no hard-link support), the export fails closed with a clear
    Greek error and no leftover temp files.
    """
    parent = os.path.dirname(path) or "."
    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", dir=parent)
    try:
        os.close(tmp_fd)
        wb.save(tmp_path)
        os.link(tmp_path, path)
    except OSError:
        # FileExistsError (dst exists), EXDEV (cross-device link),
        # ENOTSUP (filesystem doesn't support hard links), or any
        # other OS-level failure.  The target is never ours to touch.
        _remove_if_exists(tmp_path)
        if os.path.exists(path):
            return ExportResult.failure(
                f"Το αρχείο '{os.path.basename(path)}' υπάρχει ήδη. "
                "Η αντικατάσταση υπαρχόντων αρχείων δεν επιτρέπεται."
            )
        return ExportResult.failure(
            "Αδυναμία εγγραφής αρχείου εξαγωγής: "
            "το σύστημα αρχείων δεν υποστηρίζει "
            "την ασφαλή ολοκλήρωση της εξαγωγής."
        )
    except Exception as exc:
        _remove_if_exists(tmp_path)
        return ExportResult.failure(
            "Αδυναμία εγγραφής αρχείου εξαγωγής: "
            f"σφάλμα κατά την αποθήκευση — {exc}"
        )

    # Success — unlink the temp directory entry; data lives at *path*.
    try:
        os.unlink(tmp_path)
    except OSError:
        # Clean up both entries on unlink failure to avoid orphans.
        _remove_if_exists(path)
        _remove_if_exists(tmp_path)
        return ExportResult.failure(
            "Αδυναμία εγγραφής αρχείου εξαγωγής: "
            "σφάλμα κατά την αφαίρεση προσωρινού αρχείου."
        )
    return ExportResult.success(path)


def _remove_if_exists(p: str) -> None:
    """Best-effort removal of a single path — never raises."""
    try:
        if os.path.exists(p):
            os.remove(p)
    except OSError:
        pass


# ── Public export API ─────────────────────────────────────────────────

def export_inventory_snapshot(
    snapshot: "InventorySnapshot",
    target_path: str,
) -> ExportResult:
    """Export an ``InventorySnapshot`` to a single-worksheet .xlsx file.

    Worksheet: ``Αποθήκη``
    Columns:  Barcode, Προϊόν, Απόθεμα, Ημ/νία Λήξης, Τιμή,
              Προμηθευτής, Κατάσταση
    """
    err = _validate_xlsx_path(target_path)
    if err is not None:
        return err

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Αποθήκη"

        headers = [
            "Barcode", "Προϊόν", "Απόθεμα", "Ημ/νία Λήξης",
            "Τιμή", "Προμηθευτής", "Κατάσταση",
        ]
        _write_header(ws, headers)

        for row_idx, product in enumerate(snapshot.products, start=2):
            ws.cell(row=row_idx, column=1, value=product.barcode)
            ws.cell(row=row_idx, column=2, value=product.name)
            ws.cell(row=row_idx, column=3, value=product.stock)
            ws.cell(row=row_idx, column=4, value=product.expiry_date)
            ws.cell(row=row_idx, column=5, value=product.price)
            ws.cell(row=row_idx, column=6, value=product.supplier_name)
            ws.cell(row=row_idx, column=7,
                    value=", ".join(product.status_labels))

        _freeze_header(ws)
        _auto_width(ws)

        return _safe_write(target_path, wb)
    except Exception as exc:
        return ExportResult.failure(
            "Αδυναμία εξαγωγής αποθήκης: "
            f"σφάλμα κατά τη δημιουργία αρχείου — {exc}"
        )


def export_daily_alerts(
    snapshot: "DailyAlertsSnapshot",
    target_path: str,
) -> ExportResult:
    """Export a ``DailyAlertsSnapshot`` to a single-worksheet .xlsx file.

    Worksheet: ``Ειδοποιήσεις``
    Columns:  Barcode, Προϊόν, Απόθεμα, Ημ/νία Λήξης, Τιμή,
              Λόγοι Ειδοποίησης
    """
    err = _validate_xlsx_path(target_path)
    if err is not None:
        return err

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ειδοποιήσεις"

        headers = [
            "Barcode", "Προϊόν", "Απόθεμα", "Ημ/νία Λήξης",
            "Τιμή", "Λόγοι Ειδοποίησης",
        ]
        _write_header(ws, headers)

        for row_idx, item in enumerate(snapshot.items, start=2):
            ws.cell(row=row_idx, column=1, value=item.barcode)
            ws.cell(row=row_idx, column=2, value=item.name)
            ws.cell(row=row_idx, column=3, value=item.stock)
            ws.cell(row=row_idx, column=4, value=item.expiry_date)
            ws.cell(row=row_idx, column=5, value=item.price)
            ws.cell(row=row_idx, column=6,
                    value=", ".join(item.reasons))

        _freeze_header(ws)
        _auto_width(ws)

        return _safe_write(target_path, wb)
    except Exception as exc:
        return ExportResult.failure(
            "Αδυναμία εξαγωγής ειδοποιήσεων: "
            f"σφάλμα κατά τη δημιουργία αρχείου — {exc}"
        )


def export_supplier_reorder(
    result: "SupplierReorderResult",
    target_path: str,
) -> ExportResult:
    """Export ``SupplierReorderResult`` to a multi-worksheet .xlsx file.

    Worksheet 1: ``Υποψήφιοι Αναπαραγγελίας``
        Single frozen header row with supplier name carried as a
        data column.  Blank separator rows between supplier groups.

    Worksheet 2: ``Προϊόντα Χωρίς Προμηθευτή``
        Unassigned products with a Greek reason column.

    Columns (assigned):  Barcode, Προϊόν, Απόθεμα, Όριο,
                          Ημ/νία Λήξης, Τιμή, Προμηθευτής
    Columns (unassigned): Barcode, Προϊόν, Απόθεμα, Όριο,
                           Ημ/νία Λήξης, Τιμή, Λόγος
    """
    err = _validate_xlsx_path(target_path)
    if err is not None:
        return err

    try:
        wb = Workbook()

        # ── Sheet 1: Candidates by supplier ───────────────────────────
        ws1 = wb.active
        ws1.title = "Υποψήφιοι Αναπαραγγελίας"

        product_headers = [
            "Barcode", "Προϊόν", "Απόθεμα", "Όριο",
            "Ημ/νία Λήξης", "Τιμή", "Προμηθευτής",
        ]
        _write_header(ws1, product_headers)

        row = 2
        for gi, group in enumerate(result.groups):
            if gi > 0:
                row += 1  # blank separator between supplier groups

            for product in group.products:
                ws1.cell(row=row, column=1, value=product.barcode)
                ws1.cell(row=row, column=2, value=product.name)
                ws1.cell(row=row, column=3, value=product.stock)
                ws1.cell(row=row, column=4, value=product.threshold)
                ws1.cell(row=row, column=5, value=product.expiry_date)
                ws1.cell(row=row, column=6, value=product.price)
                ws1.cell(row=row, column=7,
                         value=group.supplier_name)
                row += 1

        _freeze_header(ws1)
        _auto_width(ws1)

        # ── Sheet 2: Unassigned products ──────────────────────────────
        ws2 = wb.create_sheet(title="Προϊόντα Χωρίς Προμηθευτή")

        unassigned_headers = [
            "Barcode", "Προϊόν", "Απόθεμα", "Όριο",
            "Ημ/νία Λήξης", "Τιμή", "Λόγος",
        ]
        _write_header(ws2, unassigned_headers)

        for row_idx, item in enumerate(result.unassigned, start=2):
            ws2.cell(row=row_idx, column=1, value=item.barcode)
            ws2.cell(row=row_idx, column=2, value=item.name)
            ws2.cell(row=row_idx, column=3, value=item.stock)
            ws2.cell(row=row_idx, column=4, value=item.threshold)
            ws2.cell(row=row_idx, column=5, value=item.expiry_date)
            ws2.cell(row=row_idx, column=6, value=item.price)
            ws2.cell(row=row_idx, column=7, value=item.reason)

        _freeze_header(ws2)
        _auto_width(ws2)

        return _safe_write(target_path, wb)
    except Exception as exc:
        return ExportResult.failure(
            "Αδυναμία εξαγωγής υποψηφίων αναπαραγγελίας: "
            f"σφάλμα κατά τη δημιουργία αρχείου — {exc}"
        )
