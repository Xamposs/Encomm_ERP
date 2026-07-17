"""
Atomic update of changed existing products from XLSX (Phase C3).

Updates Name, Stock, Price, and ExpiryDate for products classified as
changed during review.  Every write is fully atomic — any failure,
cancellation, stale source, stale database state, invariant violation,
or audit failure rolls back the entire operation.

Never touches Barcode, supplier data, barcode metadata, EOF, schema,
or VAT-related fields.
"""

from __future__ import annotations

import hashlib
import os
import sqlite3
from dataclasses import dataclass
from decimal import Decimal

import openpyxl
from infrastructure.product_import_preview import ImportColumnMapping
from infrastructure.product_import_conflicts import (
    _normalize_barcode, _normalize_name, _normalize_stock,
    _normalize_price, _normalize_expiry, IncomingProduct,
    _compute_review_db_signature,
)
from infrastructure.import_constants import MAX_IMPORT_ROWS
from infrastructure.product_import_plan import ImportPlan
from infrastructure.product_import_identity import (
    verify_import_source, ImportSourceSignature,
)
from infrastructure.database_service import DatabaseService


@dataclass
class ImportUpdateCommitResult:
    """Typed result for C3 atomic update operations."""
    ok: bool
    cancelled: bool
    error_message: str
    updated_rows: int
    updated_name: int
    updated_stock: int
    updated_price: int
    updated_expiry: int
    skipped_identical: int
    skipped_new: int
    rejected_invalid: int
    skipped_duplicates: int
    source_signature: ImportSourceSignature | None
    review_db_signature: str | None

    @classmethod
    def success(cls, updated_rows, updated_name, updated_stock,
                updated_price, updated_expiry, skipped_identical,
                skipped_new, rejected, skipped_dup, sig, db_sig):
        return cls(True, False, "", updated_rows, updated_name,
                   updated_stock, updated_price, updated_expiry,
                   skipped_identical, skipped_new, rejected,
                   skipped_dup, sig, db_sig)

    @classmethod
    def cancelled_result(cls, sig=None, db_sig=None):
        return cls(False, True, "Η ενημέρωση ακυρώθηκε.",
                   0, 0, 0, 0, 0, 0, 0, 0, 0, sig, db_sig)

    @classmethod
    def failure(cls, msg, sig=None, db_sig=None):
        return cls(False, False, msg, 0, 0, 0, 0, 0, 0, 0, 0, 0,
                   sig, db_sig)


def _connect_rw(db_path: str) -> sqlite3.Connection:
    if not os.path.isfile(db_path):
        raise FileNotFoundError(f"Η βάση δεδομένων δεν βρέθηκε: {db_path}")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _verify_productmaster_columns(conn):
    info = conn.execute(
        "SELECT name FROM pragma_table_info('ProductMaster')").fetchall()
    cols = {r["name"] for r in info}
    for required in ["Barcode", "Name", "Stock", "Price", "ExpiryDate"]:
        if required not in cols:
            raise ValueError(f"Λείπει η στήλη {required} από τον ProductMaster.")


def _revalidate_stream(file_path, mapping, sheet_name, conn,
                       cancel_event, max_rows):
    """Revalidate against current DB. Returns classification tuple.

    Matches B1/C1 validation semantics exactly.
    """
    BATCH_SIZE = 500
    valid = invalid = 0
    new_count = identical_count = changed_count = 0
    seen: set[str] = set()
    dupe_set: set[str] = set()
    incoming: dict[str, IncomingProduct] = {}
    new_barcodes: set[str] = set()
    col_map = {}

    wb = None
    try:
        wb = openpyxl.load_workbook(
            file_path, read_only=True, data_only=True, keep_links=False)
        ws = wb[sheet_name] if sheet_name else wb.active

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, values_only=True), start=1):

            if cancel_event and cancel_event.is_set():
                wb.close()
                return (False, "Ακυρώθηκε", valid, invalid, len(dupe_set),
                        0, 0, 0, set(), set())

            if row_idx == 1:
                headers = tuple(
                    str(c).strip() if c is not None else "" for c in row)
                if len(set(headers)) != len(headers):
                    wb.close()
                    return (False, "Διπλότυπα ονόματα στηλών.",
                            valid, invalid, 0, 0, 0, 0, set(), set())
                map_cols = [mapping.barcode_column, mapping.name_column,
                            mapping.stock_column, mapping.price_column,
                            mapping.expiry_date_column]
                if len(set(map_cols)) != 5:
                    wb.close()
                    return (False, "Μη μοναδική αντιστοίχιση.",
                            valid, invalid, 0, 0, 0, 0, set(), set())
                for key, col_name in [
                    ("barcode", mapping.barcode_column),
                    ("name", mapping.name_column),
                    ("stock", mapping.stock_column),
                    ("price", mapping.price_column),
                    ("expiry", mapping.expiry_date_column)]:
                    try:
                        col_map[key] = headers.index(col_name)
                    except ValueError:
                        wb.close()
                        return (False,
                                f"Στήλη '{col_name}' δεν βρέθηκε.",
                                valid, invalid, 0, 0, 0, 0, set(), set())
                continue

            if row_idx - 1 >= max_rows:
                wb.close()
                return (False,
                        f"Υπέρβαση ορίου {max_rows:,} γραμμών.",
                        valid, invalid, len(dupe_set), 0, 0, 0, set(), set())

            def _col(key):
                idx = col_map[key]
                return row[idx] if idx < len(row) else None

            barcode = _normalize_barcode(_col("barcode"))
            name = _normalize_name(_col("name"))
            stock = _normalize_stock(_col("stock"))
            price = _normalize_price(_col("price"))
            expiry = _normalize_expiry(_col("expiry"))

            if not all([barcode, name, stock is not None,
                        price is not None, expiry is not None]):
                invalid += 1
                continue
            if barcode in seen:
                dupe_set.add(barcode)
                invalid += 1
                continue
            seen.add(barcode)
            valid += 1

            prod = IncomingProduct(barcode, name, stock, price, expiry or "")
            incoming[barcode] = prod

            if len(incoming) >= BATCH_SIZE:
                nc, ic, cc, nbc = _classify_chunk(
                    incoming, conn, cancel_event)
                new_count += nc
                identical_count += ic
                changed_count += cc
                new_barcodes |= nbc
                incoming.clear()

        if incoming:
            nc, ic, cc, nbc = _classify_chunk(incoming, conn, cancel_event)
            new_count += nc
            identical_count += ic
            changed_count += cc
            new_barcodes |= nbc

        return (True, "", valid, invalid, len(dupe_set),
                new_count, identical_count, changed_count, new_barcodes,
                seen)

    except Exception as e:
        return (False, str(e), 0, 0, 0, 0, 0, 0, set(), set())
    finally:
        if wb:
            wb.close()


def _classify_chunk(incoming, conn, cancel_event):
    """Returns (new, identical, changed, new_barcodes)."""
    barcodes = list(incoming.keys())
    placeholders = ",".join("?" for _ in barcodes)
    rows = conn.execute(
        f"SELECT Barcode, Name, Stock, Price, ExpiryDate "
        f"FROM ProductMaster WHERE Barcode IN ({placeholders})",
        barcodes).fetchall()
    db_map = {r["Barcode"]: r for r in rows}

    new = identical = changed = 0
    new_bcs: set[str] = set()
    for barcode in barcodes:
        if cancel_event and cancel_event.is_set():
            return new, identical, changed, new_bcs
        prod = incoming[barcode]
        existing = db_map.get(barcode)
        if existing is None:
            new += 1
            new_bcs.add(barcode)
            continue
        name_match = prod.name == (existing["Name"] or "")
        stock_match = prod.stock == (existing["Stock"] or 0)
        price_match = (Decimal(str(prod.price))
                       == Decimal(str(existing["Price"] or 0.0)))
        expiry_match = ((prod.expiry_date or "")
                        == (existing["ExpiryDate"] or ""))
        if name_match and stock_match and price_match and expiry_match:
            identical += 1
        else:
            changed += 1
    return new, identical, changed, new_bcs


def commit_approved_changed_products_from_xlsx(
    file_path: str,
    mapping: ImportColumnMapping,
    plan: ImportPlan,
    db_path: str,
    cancel_event=None,
) -> ImportUpdateCommitResult:
    """Atomically update existing products that were reviewed and approved.

    All validation, revalidation, and writes happen inside a single
    BEGIN IMMEDIATE transaction.  Any failure rolls back every write.
    """

    # ── Gate 1: plan prerequisites ──────────────────────────────────
    if not plan.read_only:
        return ImportUpdateCommitResult.failure(
            "Το σχέδιο δεν είναι ανάγνωσης μόνο.")
    if plan.source_signature is None:
        return ImportUpdateCommitResult.failure(
            "Το σχέδιο δεν έχει ταυτότητα αρχείου.")
    if plan.review_db_signature is None:
        return ImportUpdateCommitResult.failure(
            "Το σχέδιο δεν έχει υπογραφή βάσης (review_db_signature). "
            "Δημιουργήστε νέο σχέδιο.")
    if plan.skipped_changed != 0:
        return ImportUpdateCommitResult.failure(
            "Το σχέδιο έχει παραλειφθείσες αλλαγές "
            "(skipped_changed != 0). "
            "Απαιτείται σχέδιο με χειροκίνητο έλεγχο.")
    if plan.manual_review == 0:
        return ImportUpdateCommitResult.failure(
            "Δεν υπάρχουν προϊόντα προς ενημέρωση (manual_review=0).")
    if not os.path.isfile(db_path):
        return ImportUpdateCommitResult.failure(
            f"Η βάση δεδομένων δεν βρέθηκε: {db_path}")

    # ── Gate 2: source signature (before transaction) ───────────────
    if not verify_import_source(plan.source_signature, file_path, mapping):
        return ImportUpdateCommitResult.failure(
            "Το αρχείο ή η αντιστοίχιση άλλαξε από τη δημιουργία "
            "του σχεδίου. Δημιουργήστε νέο σχέδιο.")

    conn = None
    try:
        conn = _connect_rw(db_path)
        conn.execute("BEGIN IMMEDIATE")
        _verify_productmaster_columns(conn)

        # ── Gate 3: source signature (inside transaction) ───────────
        if not verify_import_source(plan.source_signature,
                                     file_path, mapping):
            conn.rollback()
            conn.close()
            return ImportUpdateCommitResult.failure(
                "Το αρχείο άλλαξε μετά το κλείδωμα. "
                "Δημιουργήστε νέο σχέδιο.")

        # ── Revalidate stream (B1/C1 rules, inside txn) ─────────────
        rv_ok, rv_err, valid, invalid, dupes, new_cnt, ident_cnt, \
            chg_cnt, new_bcs, seen_barcodes = _revalidate_stream(
                file_path, mapping, plan.sheet_name, conn,
                cancel_event, MAX_IMPORT_ROWS)

        if cancel_event and cancel_event.is_set():
            conn.rollback()
            conn.close()
            return ImportUpdateCommitResult.cancelled_result(
                plan.source_signature, plan.review_db_signature)

        if not rv_ok:
            conn.rollback()
            conn.close()
            return ImportUpdateCommitResult.failure(
                f"Αποτυχία επανεπικύρωσης: {rv_err}")

        changed_total = plan.manual_review + plan.skipped_changed

        # ── Plan invariant validation ───────────────────────────────
        invariants = [
            (valid == plan.valid_rows, "valid_rows"),
            (plan.classified_rows == plan.valid_rows,
             "classified_rows == valid_rows"),
            (plan.classified_rows == plan.planned_new
             + plan.skipped_identical + changed_total,
             "classified_rows sum"),
            (invalid == plan.invalid_rows, "invalid_rows"),
            (invalid == plan.rejected_invalid,
             "rejected_invalid == invalid_rows"),
            (dupes == plan.duplicate_barcodes, "duplicate_barcodes"),
            (dupes == plan.skipped_duplicates,
             "skipped_duplicates == duplicate_barcodes"),
            (new_cnt == plan.planned_new, "planned_new"),
            (ident_cnt == plan.skipped_identical, "skipped_identical"),
            (chg_cnt == changed_total, "changed_total"),
        ]
        for passes, label in invariants:
            if not passes:
                conn.rollback()
                conn.close()
                return ImportUpdateCommitResult.failure(
                    f"Ασυνέπεια σχεδίου ({label}). "
                    f"Δημιουργήστε νέο σχέδιο.")

        # ── Recalculate DB review signature (inside txn) ────────────
        recalc_sig = _compute_review_db_signature(conn, seen_barcodes)
        if recalc_sig != plan.review_db_signature:
            conn.rollback()
            conn.close()
            return ImportUpdateCommitResult.failure(
                "Η βάση δεδομένων άλλαξε από τη στιγμή του ελέγχου. "
                "Δημιουργήστε νέα ανάλυση και νέο σχέδιο.")

        # ── Streamed UPDATE pass ────────────────────────────────────
        updated_rows = 0
        updated_name = 0
        updated_stock = 0
        updated_price = 0
        updated_expiry = 0

        wb = None
        try:
            wb = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True, keep_links=False)
            ws = wb[plan.sheet_name] if plan.sheet_name else wb.active
            col_map: dict[str, int] = {}
            seen_update: set[str] = set()

            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, values_only=True), start=1):

                if cancel_event and cancel_event.is_set():
                    conn.rollback()
                    if wb:
                        wb.close()
                    conn.close()
                    return ImportUpdateCommitResult.cancelled_result(
                        plan.source_signature, plan.review_db_signature)

                if row_idx == 1:
                    headers = tuple(
                        str(c).strip() if c is not None else "" for c in row)
                    for key, col_name in [
                        ("barcode", mapping.barcode_column),
                        ("name", mapping.name_column),
                        ("stock", mapping.stock_column),
                        ("price", mapping.price_column),
                        ("expiry", mapping.expiry_date_column)]:
                        try:
                            col_map[key] = headers.index(col_name)
                        except ValueError:
                            conn.rollback()
                            if wb:
                                wb.close()
                            conn.close()
                            return ImportUpdateCommitResult.failure(
                                f"Στήλη '{col_name}' δεν βρέθηκε.")
                    continue

                if row_idx - 1 >= MAX_IMPORT_ROWS:
                    conn.rollback()
                    if wb:
                        wb.close()
                    conn.close()
                    return ImportUpdateCommitResult.failure(
                        f"Υπέρβαση ορίου {MAX_IMPORT_ROWS:,} γραμμών.")

                def _col(key):
                    idx = col_map[key]
                    return row[idx] if idx < len(row) else None

                barcode = _normalize_barcode(_col("barcode"))
                name = _normalize_name(_col("name"))
                stock = _normalize_stock(_col("stock"))
                price = _normalize_price(_col("price"))
                expiry = _normalize_expiry(_col("expiry"))

                # Only process barcodes that revalidation classified as
                # changed (not new, not identical, not invalid)
                if barcode is None:
                    continue
                if barcode in new_bcs:
                    # New product — C3 never touches these
                    continue
                if barcode in seen_update:
                    continue

                # Check: must exist in DB and be classified as changed
                existing = conn.execute(
                    "SELECT Barcode, Name, Stock, Price, ExpiryDate "
                    "FROM ProductMaster WHERE Barcode=?",
                    (barcode,)).fetchone()
                if existing is None:
                    # Not in DB → was new during revalidation, skip
                    continue

                # Classify this row against current DB to confirm it's changed
                name_match = (name or "") == (existing["Name"] or "")
                stock_match = (stock or 0) == (existing["Stock"] or 0)
                price_match = (Decimal(str(price))
                               == Decimal(str(existing["Price"] or 0.0)))
                expiry_match = ((expiry or "")
                                == (existing["ExpiryDate"] or ""))

                if name_match and stock_match and price_match and expiry_match:
                    # Identical — C3 never updates these
                    continue

                seen_update.add(barcode)

                # ── Parameterized UPDATE — only Name, Stock, Price, ExpiryDate ──
                conn.execute(
                    "UPDATE ProductMaster "
                    "SET Name=?, Stock=?, Price=?, ExpiryDate=? "
                    "WHERE Barcode=?",
                    (name, stock, float(price), expiry or "", barcode))

                # ── Count per-field changes ──
                if not name_match:
                    updated_name += 1
                if not stock_match:
                    updated_stock += 1
                if not price_match:
                    updated_price += 1
                if not expiry_match:
                    updated_expiry += 1

                # ── Stock movement audit (only when stock changes) ──
                if not stock_match:
                    old_st = existing["Stock"] or 0
                    DatabaseService._log_stock_movement_on_conn(
                        conn, barcode, name or existing["Name"] or "",
                        old_stock=old_st,
                        new_stock=stock,
                        reason="Ενημέρωση Excel (εγκεκριμένη)",
                        source="Excel Import",
                        operator="Σύστημα",
                    )

                updated_rows += 1

        finally:
            if wb:
                wb.close()

        # ── Verification gate: updated_rows must match plan ─────────
        if updated_rows != plan.manual_review:
            conn.rollback()
            conn.close()
            return ImportUpdateCommitResult.failure(
                f"Ασυνέπεια: ενημερώθηκαν {updated_rows} αντί για "
                f"{plan.manual_review}. Δημιουργήστε νέο σχέδιο.")

        # ── Gate 4: source signature (before commit) ────────────────
        if not verify_import_source(plan.source_signature, file_path,
                                     mapping):
            conn.rollback()
            conn.close()
            return ImportUpdateCommitResult.failure(
                "Το αρχείο άλλαξε πριν την ολοκλήρωση. "
                "Δημιουργήστε νέο σχέδιο.")

        conn.commit()
        conn.close()

        return ImportUpdateCommitResult.success(
            updated_rows, updated_name, updated_stock,
            updated_price, updated_expiry,
            plan.skipped_identical, plan.planned_new,
            plan.rejected_invalid, plan.skipped_duplicates,
            plan.source_signature, plan.review_db_signature)

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
            conn.close()
        return ImportUpdateCommitResult.failure(
            f"Σφάλμα ενημέρωσης: {e}")
