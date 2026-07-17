"""
Atomic commit of NEW products from XLSX (Phase C1).

Insert-only — never updates, overwrites, or deletes existing products.
"""

from __future__ import annotations

import os
import sqlite3
from dataclasses import dataclass
from decimal import Decimal

import openpyxl
from infrastructure.product_import_preview import ImportColumnMapping
from infrastructure.product_import_conflicts import (
    _normalize_barcode, _normalize_name, _normalize_stock,
    _normalize_price, _normalize_expiry, IncomingProduct,
)
from infrastructure.import_constants import MAX_IMPORT_ROWS
from infrastructure.product_import_plan import ImportPlan
from infrastructure.product_import_identity import (
    verify_import_source, ImportSourceSignature,
)
from infrastructure.database_service import DatabaseService


@dataclass
class ImportCommitResult:
    ok: bool
    cancelled: bool
    error_message: str
    inserted_rows: int
    skipped_identical: int
    skipped_changed: int
    rejected_invalid: int
    skipped_duplicates: int
    source_signature: ImportSourceSignature | None

    @classmethod
    def success(cls, inserted, skipped_id, skipped_ch, rejected, skipped_dup,
                sig):
        return cls(True, False, "", inserted, skipped_id, skipped_ch,
                   rejected, skipped_dup, sig)

    @classmethod
    def cancelled_result(cls, sig):
        return cls(False, True, "Η εισαγωγή ακυρώθηκε.",
                   0, 0, 0, 0, 0, sig)

    @classmethod
    def failure(cls, msg, sig=None):
        return cls(False, False, msg, 0, 0, 0, 0, 0, sig)


def _connect_rw(db_path: str) -> sqlite3.Connection:
    if not os.path.isfile(db_path):
        raise FileNotFoundError(f"Η βάση δεδομένων δεν βρέθηκε: {db_path}")
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def _verify_productmaster_columns(conn):
    info = conn.execute(
        "SELECT name FROM pragma_table_info('ProductMaster')").fetchall()
    cols = {r["name"] for r in info}
    for required in ["Barcode", "Name", "Stock", "Price", "ExpiryDate"]:
        if required not in cols:
            raise ValueError(f"Λείπει η στήλη {required} από τον ProductMaster.")


def _revalidate_stream(
    file_path, mapping, sheet_name, conn, cancel_event, max_rows,
):
    """Revalidate against current DB.  Returns tuple.
    Matches B1 validation semantics exactly."""
    BATCH_SIZE = 500
    valid = invalid = 0
    new_count = identical_count = changed_count = 0
    seen: set[str] = set()
    dupe_set: set[str] = set()
    incoming: dict[str, IncomingProduct] = {}
    new_barcodes: set[str] = set()
    col_map = {}

    wb = None
    try:
        wb = openpyxl.load_workbook(
            file_path, read_only=True, data_only=True, keep_links=False)
        ws = wb[sheet_name] if sheet_name else wb.active

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, values_only=True), start=1):

            if cancel_event and cancel_event.is_set():
                wb.close()
                return (False, "Ακυρώθηκε", valid, invalid, len(dupe_set),
                        0, 0, 0, set())

            if row_idx == 1:
                headers = tuple(
                    str(c).strip() if c is not None else "" for c in row)
                if len(set(headers)) != len(headers):
                    wb.close()
                    return (False, "Διπλότυπα ονόματα στηλών.",
                            valid, invalid, 0, 0, 0, 0, set())
                map_cols = [mapping.barcode_column, mapping.name_column,
                            mapping.stock_column, mapping.price_column,
                            mapping.expiry_date_column]
                if len(set(map_cols)) != 5:
                    wb.close()
                    return (False, "Μη μοναδική αντιστοίχιση.",
                            valid, invalid, 0, 0, 0, 0, set())
                for key, col_name in [
                    ("barcode", mapping.barcode_column),
                    ("name", mapping.name_column),
                    ("stock", mapping.stock_column),
                    ("price", mapping.price_column),
                    ("expiry", mapping.expiry_date_column)]:
                    try:
                        col_map[key] = headers.index(col_name)
                    except ValueError:
                        wb.close()
                        return (False,
                                f"Στήλη '{col_name}' δεν βρέθηκε.",
                                valid, invalid, 0, 0, 0, 0, set())
                continue

            if row_idx - 1 >= max_rows:
                wb.close()
                return (False,
                        f"Υπέρβαση ορίου {max_rows:,} γραμμών.",
                        valid, invalid, len(dupe_set), 0, 0, 0, set())

            def _col(key):
                idx = col_map[key]
                return row[idx] if idx < len(row) else None

            # Use B1 normalization exactly
            barcode = _normalize_barcode(_col("barcode"))
            name = _normalize_name(_col("name"))
            stock = _normalize_stock(_col("stock"))
            price = _normalize_price(_col("price"))
            expiry = _normalize_expiry(_col("expiry"))

            if not all([barcode, name, stock is not None,
                        price is not None, expiry is not None]):
                invalid += 1
                continue
            if barcode in seen:
                dupe_set.add(barcode)
                invalid += 1
                continue
            seen.add(barcode)
            valid += 1

            prod = IncomingProduct(
                barcode, name, stock, price, expiry or "")
            incoming[barcode] = prod

            if len(incoming) >= BATCH_SIZE:
                nc, ic, cc, nbc = _classify_chunk(
                    incoming, conn, cancel_event)
                new_count += nc
                identical_count += ic
                changed_count += cc
                new_barcodes |= nbc
                incoming.clear()

        if incoming:
            nc, ic, cc, nbc = _classify_chunk(
                incoming, conn, cancel_event)
            new_count += nc
            identical_count += ic
            changed_count += cc
            new_barcodes |= nbc

        return (True, "", valid, invalid, len(dupe_set),
                new_count, identical_count, changed_count, new_barcodes)

    except Exception as e:
        return (False, str(e), 0, 0, 0, 0, 0, 0, set())
    finally:
        if wb:
            wb.close()


def _classify_chunk(incoming, conn, cancel_event):
    """Returns (new, identical, changed, new_barcodes)."""
    barcodes = list(incoming.keys())
    placeholders = ",".join("?" for _ in barcodes)
    rows = conn.execute(
        f"SELECT Barcode, Name, Stock, Price, ExpiryDate "
        f"FROM ProductMaster WHERE Barcode IN ({placeholders})",
        barcodes).fetchall()
    db_map = {r["Barcode"]: r for r in rows}

    new = identical = changed = 0
    new_bcs: set[str] = set()
    for barcode in barcodes:
        if cancel_event and cancel_event.is_set():
            return new, identical, changed, new_bcs
        prod = incoming[barcode]
        existing = db_map.get(barcode)
        if existing is None:
            new += 1
            new_bcs.add(barcode)
            continue
        # Exact Decimal price comparison — matches B1 semantics
        name_match = prod.name == (existing["Name"] or "")
        stock_match = prod.stock == (existing["Stock"] or 0)
        price_match = (Decimal(str(prod.price))
                       == Decimal(str(existing["Price"] or 0.0)))
        expiry_match = ((prod.expiry_date or "")
                        == (existing["ExpiryDate"] or ""))
        if name_match and stock_match and price_match and expiry_match:
            identical += 1
        else:
            changed += 1
    return new, identical, changed, new_bcs


def commit_new_products_from_xlsx(
    file_path: str,
    mapping: ImportColumnMapping,
    plan: ImportPlan,
    db_path: str,
    cancel_event=None,
) -> ImportCommitResult:
    if not plan.read_only:
        return ImportCommitResult.failure(
            "Το σχέδιο δεν είναι ανάγνωσης μόνο.")
    if plan.source_signature is None:
        return ImportCommitResult.failure(
            "Το σχέδιο δεν έχει ταυτότητα αρχείου.")
    if plan.planned_new == 0:
        return ImportCommitResult.failure(
            "Δεν υπάρχουν νέα προϊόντα για εισαγωγή.")
    # Check DB exists before signature (avoid confusing error msg)
    if not os.path.isfile(db_path):
        return ImportCommitResult.failure(
            f"Η βάση δεδομένων δεν βρέθηκε: {db_path}")
    if not verify_import_source(plan.source_signature, file_path, mapping):
        return ImportCommitResult.failure(
            "Το αρχείο ή η αντιστοίχιση άλλαξε από τη δημιουργία "
            "του σχεδίου. Δημιουργήστε νέο σχέδιο.")

    conn = None
    try:
        conn = _connect_rw(db_path)
        conn.execute("BEGIN IMMEDIATE")
        _verify_productmaster_columns(conn)

        if not verify_import_source(plan.source_signature, file_path, mapping):
            conn.rollback()
            conn.close()
            return ImportCommitResult.failure(
                "Το αρχείο άλλαξε μετά το κλείδωμα. "
                "Δημιουργήστε νέο σχέδιο.")

        ok, err, valid, invalid, dupes, new_cnt, ident_cnt, chg_cnt, new_bcs = (
            _revalidate_stream(file_path, mapping, plan.sheet_name,
                               conn, cancel_event, MAX_IMPORT_ROWS))

        if cancel_event and cancel_event.is_set():
            conn.rollback()
            conn.close()
            return ImportCommitResult.cancelled_result(plan.source_signature)

        if not ok:
            conn.rollback()
            conn.close()
            return ImportCommitResult.failure(
                f"Αποτυχία επανεπικύρωσης: {err}")

        changed_total = plan.manual_review + plan.skipped_changed

        # ── Plan invariant validation ──────────────────────────────
        invariants = [
            (valid == plan.valid_rows,
             "valid_rows"),
            (plan.classified_rows == plan.valid_rows,
             "classified_rows == valid_rows"),
            (plan.classified_rows == plan.planned_new
             + plan.skipped_identical + changed_total,
             "classified_rows sum"),
            (invalid == plan.invalid_rows,
             "invalid_rows"),
            (invalid == plan.rejected_invalid,
             "rejected_invalid == invalid_rows"),
            (dupes == plan.duplicate_barcodes,
             "duplicate_barcodes"),
            (dupes == plan.skipped_duplicates,
             "skipped_duplicates == duplicate_barcodes"),
            (new_cnt == plan.planned_new,
             "planned_new"),
            (ident_cnt == plan.skipped_identical,
             "skipped_identical"),
            (chg_cnt == changed_total,
             "changed_total"),
        ]

        for passes, label in invariants:
            if not passes:
                conn.rollback()
                conn.close()
                return ImportCommitResult.failure(
                    f"Ασυνέπεια σχεδίου ({label}). "
                    f"Δημιουργήστε νέο σχέδιο.")

        # ── Insert pass ────────────────────────────────────────────
        inserted = 0
        wb = None
        try:
            wb = openpyxl.load_workbook(
                file_path, read_only=True, data_only=True, keep_links=False)
            ws = wb[plan.sheet_name] if plan.sheet_name else wb.active
            col_map: dict[str, int] = {}
            seen_insert: set[str] = set()

            for row_idx, row in enumerate(
                ws.iter_rows(min_row=1, values_only=True), start=1):

                if cancel_event and cancel_event.is_set():
                    conn.rollback()
                    if wb:
                        wb.close()
                    conn.close()
                    return ImportCommitResult.cancelled_result(
                        plan.source_signature)

                if row_idx == 1:
                    headers = tuple(
                        str(c).strip() if c is not None else "" for c in row)
                    for key, col_name in [
                        ("barcode", mapping.barcode_column),
                        ("name", mapping.name_column),
                        ("stock", mapping.stock_column),
                        ("price", mapping.price_column),
                        ("expiry", mapping.expiry_date_column)]:
                        try:
                            col_map[key] = headers.index(col_name)
                        except ValueError:
                            conn.rollback()
                            if wb:
                                wb.close()
                            conn.close()
                            return ImportCommitResult.failure(
                                f"Στήλη '{col_name}' δεν βρέθηκε.")
                    continue

                if row_idx - 1 >= MAX_IMPORT_ROWS:
                    conn.rollback()
                    if wb:
                        wb.close()
                    conn.close()
                    return ImportCommitResult.failure(
                        f"Υπέρβαση ορίου {MAX_IMPORT_ROWS:,} γραμμών.")

                def _col(key):
                    idx = col_map[key]
                    return row[idx] if idx < len(row) else None

                barcode = _normalize_barcode(_col("barcode"))
                name = _normalize_name(_col("name"))
                stock = _normalize_stock(_col("stock"))
                price = _normalize_price(_col("price"))
                expiry = _normalize_expiry(_col("expiry"))

                # Only insert barcodes that revalidation classified as new
                if barcode is None or barcode not in new_bcs:
                    continue
                if barcode in seen_insert:
                    continue
                seen_insert.add(barcode)

                existing = conn.execute(
                    "SELECT Barcode FROM ProductMaster WHERE Barcode=?",
                    (barcode,)).fetchone()
                if existing is not None:
                    continue

                conn.execute(
                    "INSERT INTO ProductMaster "
                    "(Barcode, Name, Stock, ExpiryDate, Price) "
                    "VALUES (?,?,?,?,?)",
                    (barcode, name, stock, expiry, price))

                DatabaseService._log_stock_movement_on_conn(
                    conn, barcode, name,
                    old_stock=0, new_stock=stock,
                    reason="Εισαγωγή Excel",
                    source="Excel Import",
                    operator="Σύστημα")

                inserted += 1

        finally:
            if wb:
                wb.close()

        if inserted != plan.planned_new:
            conn.rollback()
            conn.close()
            return ImportCommitResult.failure(
                f"Ασυνέπεια: εισήχθησαν {inserted} αντί για "
                f"{plan.planned_new}. Δημιουργήστε νέο σχέδιο.")

        if not verify_import_source(plan.source_signature, file_path, mapping):
            conn.rollback()
            conn.close()
            return ImportCommitResult.failure(
                "Το αρχείο άλλαξε πριν την ολοκλήρωση. "
                "Δημιουργήστε νέο σχέδιο.")

        conn.commit()
        conn.close()

        return ImportCommitResult.success(
            inserted, plan.skipped_identical,
            plan.manual_review + plan.skipped_changed,
            plan.rejected_invalid, plan.skipped_duplicates,
            plan.source_signature)

    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except Exception:
                pass
            conn.close()
        return ImportCommitResult.failure(f"Σφάλμα εισαγωγής: {e}")
