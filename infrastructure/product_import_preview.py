"""
Streaming XLSX product-import preview and validation (Phase A — no writes).

Manual stress-test command (PowerShell):
  python -c "
  from infrastructure.product_import_preview import preview_product_import_xlsx, suggest_mapping
  m = suggest_mapping('large_file.xlsx')
  print(m)
  r = preview_product_import_xlsx('large_file.xlsx', m)
  print(r.valid_rows, r.invalid_rows, r.duplicate_barcodes, r.scanned_rows)
  "
"""

from __future__ import annotations

from dataclasses import dataclass, field
from typing import Tuple, Optional, List
from datetime import date, datetime

import openpyxl

# ── Public typed models ──────────────────────────────────────────────


@dataclass(frozen=True)
class ImportColumnMapping:
    barcode_column: str
    name_column: str
    stock_column: str
    price_column: str
    expiry_date_column: str


@dataclass(frozen=True)
class ImportRowError:
    row_number: int
    barcode: str
    code: str
    message: str


@dataclass(frozen=True)
class ProductImportPreview:
    ok: bool
    file_name: str
    sheet_name: str
    scanned_rows: int
    valid_rows: int
    invalid_rows: int
    duplicate_barcodes: int
    headers: Tuple[str, ...]
    detected_mapping: ImportColumnMapping | None
    sample_rows: Tuple[Tuple[str, str, int, float, str], ...] = ()
    errors: Tuple[ImportRowError, ...] = ()
    error_message: str = ""
    cancelled: bool = False

    @classmethod
    def success(cls, file_name, sheet_name, scanned, valid, invalid, dupes,
                headers, mapping, samples, errors):
        return cls(ok=True, file_name=file_name, sheet_name=sheet_name,
                   scanned_rows=scanned, valid_rows=valid,
                   invalid_rows=invalid, duplicate_barcodes=dupes,
                   headers=headers, detected_mapping=mapping,
                   sample_rows=tuple(samples),
                   errors=tuple(errors))

    @classmethod
    def failure(cls, file_name, sheet_name, msg):
        return cls(ok=False, file_name=file_name, sheet_name=sheet_name,
                   scanned_rows=0, valid_rows=0, invalid_rows=0,
                   duplicate_barcodes=0, headers=(), detected_mapping=None,
                   error_message=msg)


# ── Header aliases ───────────────────────────────────────────────────

_HEADER_ALIASES = {
    "barcode": ["barcode", "ean", "ean13", "κωδικός", "κωδικος",
                "barcode προϊόντος"],
    "name": ["name", "product", "product name", "περιγραφή", "περιγραφη",
             "όνομα", "ονομα", "προϊόν"],
    "stock": ["stock", "quantity", "qty", "απόθεμα", "αποθεμα",
              "ποσότητα", "ποσοτητα"],
    "price": ["price", "unit price", "τιμή", "τιμη", "τιμή μονάδας"],
    "expiry": ["expiry", "expiry date", "expiration", "λήξη", "ληξη",
               "ημερομηνία λήξης"],
}


def _normalize_header(val: str) -> str:
    return val.strip().lower()


# ── Public API ───────────────────────────────────────────────────────


def list_xlsx_sheets(file_path: str) -> Tuple[str, ...]:
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True,
                                keep_links=False)
    try:
        return tuple(wb.sheetnames)
    finally:
        wb.close()


def inspect_xlsx_headers(file_path: str, sheet_name: str | None = None):
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True,
                                keep_links=False)
    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(c).strip() if c is not None else "" for c in row]
            break
        return tuple(headers)
    finally:
        wb.close()


def suggest_mapping(file_path: str, sheet_name: str | None = None
                    ) -> ImportColumnMapping | None:
    """Return a suggested mapping or None if ambiguous/missing."""
    headers = inspect_xlsx_headers(file_path, sheet_name)
    mapping: dict[str, str | None] = {k: None for k in _HEADER_ALIASES}
    for idx, hdr in enumerate(headers):
        norm = _normalize_header(hdr)
        for key, aliases in _HEADER_ALIASES.items():
            if norm in [_normalize_header(a) for a in aliases]:
                if mapping[key] is not None:
                    return None  # ambiguous
                mapping[key] = hdr
    if any(v is None for v in mapping.values()):
        return None
    return ImportColumnMapping(
        barcode_column=mapping["barcode"],  # type: ignore[arg-type]
        name_column=mapping["name"],  # type: ignore[arg-type]
        stock_column=mapping["stock"],  # type: ignore[arg-type]
        price_column=mapping["price"],  # type: ignore[arg-type]
        expiry_date_column=mapping["expiry"],  # type: ignore[arg-type]
    )


def _partial_result(
    ok, cancelled, msg, file_name, sheet_name, scanned, valid, invalid,
    dupes, headers, mapping, samples, errors,
) -> ProductImportPreview:
    return ProductImportPreview(
        ok=ok, cancelled=cancelled, error_message=msg,
        file_name=file_name, sheet_name=sheet_name,
        scanned_rows=scanned, valid_rows=valid,
        invalid_rows=invalid, duplicate_barcodes=dupes,
        headers=headers, detected_mapping=mapping,
        sample_rows=tuple(samples), errors=tuple(errors),
    )


def preview_product_import_xlsx(
    file_path: str,
    mapping: ImportColumnMapping,
    sheet_name: str | None = None,
    cancel_event=None,
    max_rows: int = 250_000,
) -> ProductImportPreview:
    MAX_ERRORS = 200
    MAX_SAMPLES = 20

    # ── Mapping validation ──
    map_cols = [
        mapping.barcode_column, mapping.name_column,
        mapping.stock_column, mapping.price_column,
        mapping.expiry_date_column,
    ]
    if len(set(map_cols)) != 5:
        return ProductImportPreview.failure(
            file_path, sheet_name or "—",
            "Κάθε στήλη αντιστοίχισης πρέπει να είναι μοναδική. "
            "Δεν επιτρέπεται η ίδια στήλη για δύο διαφορετικά πεδία.")

    wb = None
    try:
        wb = openpyxl.load_workbook(
            file_path, read_only=True, data_only=True, keep_links=False)
    except Exception as e:
        return ProductImportPreview.failure(
            file_path, sheet_name or "—",
            f"Αδυναμία ανοίγματος αρχείου: {e}")

    try:
        ws = wb[sheet_name] if sheet_name else wb.active
        sheet_name = ws.title
        headers: tuple[str, ...] = ()

        col_map: dict[str, int] = {}
        seen_barcodes: set[str] = set()
        errors: list[ImportRowError] = []
        samples: list[tuple] = []
        scanned = 0
        valid = 0
        invalid = 0
        dupes = 0
        cancelled = False

        def _add_error(ri, code, msg):
            nonlocal errors
            if len(errors) < MAX_ERRORS:
                errors.append(ImportRowError(ri, "", code, msg))

        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, values_only=True), start=1
        ):
            if cancel_event and cancel_event.is_set():
                cancelled = True
                break

            if row_idx == 1:
                headers = tuple(
                    str(c).strip() if c is not None else "" for c in row)
                # Check for duplicate header names
                if len(set(headers)) != len(headers):
                    wb.close()
                    return ProductImportPreview.failure(
                        file_path, sheet_name,
                        "Η κεφαλίδα περιέχει διπλότυπα ονόματα στηλών. "
                        "Κάθε στήλη πρέπει να έχει μοναδικό όνομα.")
                for key, col_name in [
                    ("barcode", mapping.barcode_column),
                    ("name", mapping.name_column),
                    ("stock", mapping.stock_column),
                    ("price", mapping.price_column),
                    ("expiry", mapping.expiry_date_column),
                ]:
                    try:
                        col_map[key] = headers.index(col_name)
                    except ValueError:
                        wb.close()
                        return ProductImportPreview.failure(
                            file_path, sheet_name,
                            f"Η στήλη '{col_name}' δεν βρέθηκε στην κεφαλίδα. "
                            f"Διαθέσιμες στήλες: {', '.join(headers)}")
                continue

            if scanned >= max_rows:
                _add_error(row_idx, "ROWS",
                           f"Υπέρβαση ορίου {max_rows:,} γραμμών.")
                ok_result = _partial_result(
                    False, False,
                    f"Υπέρβαση ορίου {max_rows:,} γραμμών.",
                    file_path, sheet_name, scanned, valid, invalid,
                    dupes, headers, mapping, samples, errors)
                wb.close()
                return ok_result

            scanned += 1

            def _col(key):
                idx = col_map[key]
                return row[idx] if idx < len(row) else None

            barcode = _col("barcode")
            name = _col("name")
            stock_raw = _col("stock")
            price_raw = _col("price")
            expiry_raw = _col("expiry")

            row_errors: list[str] = []

            # --- barcode ---
            if barcode is None:
                barcode_str = ""
            elif isinstance(barcode, bool):
                row_errors.append("Μη έγκυρο barcode (boolean).")
                barcode_str = ""
            elif isinstance(barcode, str):
                barcode_str = barcode.strip()
            elif isinstance(barcode, (int, float)):
                from math import isfinite as _isf
                if not _isf(barcode):
                    row_errors.append("Μη έγκυρο barcode (NaN/Inf).")
                    barcode_str = ""
                elif isinstance(barcode, float) and barcode != int(barcode):
                    row_errors.append(
                        "Μη έγκυρο barcode (δεκαδικός). "
                        "Μορφοποιήστε τη στήλη ως Κείμενο στο Excel.")
                    barcode_str = ""
                else:
                    barcode_str = str(int(barcode))
                    if len(barcode_str) > 15:
                        row_errors.append(
                            "Το barcode υπερβαίνει τα 15 ψηφία. "
                            "Μορφοποιήστε τη στήλη ως Κείμενο στο Excel.")
                        barcode_str = ""
            else:
                row_errors.append("Μη έγκυρος τύπος barcode.")
                barcode_str = ""

            if not barcode_str:
                row_errors.append("Το barcode είναι κενό.")

            # --- name ---
            if name is None or (isinstance(name, str) and not name.strip()):
                row_errors.append("Το όνομα προϊόντος είναι κενό.")
                name_str = ""
            elif isinstance(name, str):
                name_str = name.strip()
            else:
                row_errors.append("Μη έγκυρος τύπος ονόματος.")
                name_str = ""

            # --- stock ---
            stock: int | None = None
            if isinstance(stock_raw, bool) or stock_raw is None:
                row_errors.append("Το απόθεμα είναι κενό ή μη έγκυρο.")
            elif isinstance(stock_raw, int):
                if stock_raw < 0:
                    row_errors.append("Το απόθεμα είναι αρνητικό.")
                else:
                    stock = stock_raw
            elif isinstance(stock_raw, float):
                from math import isfinite as _isf
                if not _isf(stock_raw):
                    row_errors.append("Μη έγκυρο απόθεμα (NaN/Inf).")
                elif stock_raw != int(stock_raw):
                    row_errors.append(
                        f"Το απόθεμα δεν είναι ακέραιο ({stock_raw}).")
                else:
                    s = int(stock_raw)
                    if s < 0:
                        row_errors.append("Το απόθεμα είναι αρνητικό.")
                    else:
                        stock = s
            else:
                row_errors.append("Μη έγκυρος τύπος αποθέματος.")

            # --- price ---
            price: float | None = None
            if isinstance(price_raw, bool) or price_raw is None:
                row_errors.append("Η τιμή είναι κενή ή μη έγκυρη.")
            elif isinstance(price_raw, (int, float)):
                p = float(price_raw)
                from math import isfinite
                if p < 0 or not isfinite(p):
                    row_errors.append(
                        "Η τιμή είναι αρνητική ή μη πεπερασμένη.")
                else:
                    price = p  # preserve exact source value
            else:
                row_errors.append("Μη έγκυρος τύπος τιμής.")

            # --- expiry ---
            expiry_str = ""
            if expiry_raw is None or (isinstance(expiry_raw, str) and not expiry_raw.strip()):
                expiry_str = ""
            elif isinstance(expiry_raw, (date, datetime)):
                expiry_str = expiry_raw.strftime("%Y-%m-%d")
            elif isinstance(expiry_raw, str):
                expiry_str = expiry_raw.strip()
                if not _is_iso_date(expiry_str):
                    row_errors.append(
                        f"Μη έγκυρη ημερομηνία λήξης: '{expiry_str}'. "
                        f"Απαιτείται YYYY-MM-DD.")
                    expiry_str = ""
            else:
                row_errors.append("Μη έγκυρος τύπος ημερομηνίας λήξης.")

            if row_errors:
                invalid += 1
                if len(errors) < MAX_ERRORS:
                    errors.append(ImportRowError(
                        row_idx, barcode_str, "VALIDATION",
                        " · ".join(row_errors)))
                continue

            # Duplicate check
            if barcode_str in seen_barcodes:
                dupes += 1
                invalid += 1
                if len(errors) < MAX_ERRORS:
                    errors.append(ImportRowError(
                        row_idx, barcode_str, "DUPLICATE",
                        f"Το barcode '{barcode_str}' εμφανίζεται ξανά στο αρχείο."))
                continue

            seen_barcodes.add(barcode_str)
            valid += 1
            if len(samples) < MAX_SAMPLES:
                samples.append(
                    (barcode_str, name_str, stock, price, expiry_str))  # type: ignore[arg-type]

        if cancelled:
            return _partial_result(
                False, True, "Η προεπισκόπηση ακυρώθηκε.",
                file_path, sheet_name, scanned, valid, invalid,
                dupes, headers, mapping, samples, errors)

        return ProductImportPreview.success(
            file_path, sheet_name, scanned, valid, invalid, dupes,
            headers, mapping, samples, errors)

    except Exception as e:
        return ProductImportPreview.failure(
            file_path, sheet_name or "—",
            f"Σφάλμα κατά την ανάγνωση: {e}")
    finally:
        if wb:
            wb.close()


def _is_iso_date(val: str) -> bool:
    try:
        date.fromisoformat(val)
        return True
    except ValueError:
        return False
